"""Excel Export utilities for sick letter records"""
import os
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import logging

logger = logging.getLogger(__name__)

BACKEND_DIR = Path(__file__).parent.parent
TEMPS_DIR = BACKEND_DIR / 'temps'

# Create temps directory if not exists
TEMPS_DIR.mkdir(exist_ok=True)


def export_records_to_excel(records: list) -> str:
    """
    Export sick letter records to Excel file
    
    Args:
        records: List of dictionaries containing patient records
    
    Returns:
        Path to the generated Excel file
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Rekap Surat Sakit"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Define headers
        headers = [
            'No',
            'Tanggal Terbit',
            'No. Surat',
            'Nama Pasien',
            'Jenis Kelamin',
            'Umur',
            'Pekerjaan',
            'Alamat',
            'Durasi Istirahat',
            'Tanggal Mulai',
            'Tanggal Selesai',
            'Diagnosa',
            'Dokter',
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Adjust column widths
        column_widths = [5, 15, 20, 25, 15, 10, 20, 30, 15, 15, 15, 30, 25]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
        
        # Write data
        for row_num, record in enumerate(records, 2):
            data_row = [
                row_num - 1,
                record.get('date_issued', ''),
                record.get('letter_number', ''),
                record.get('name', ''),
                record.get('gender', ''),
                record.get('age', ''),
                record.get('occupation', ''),
                record.get('address', ''),
                record.get('duration', ''),
                record.get('from_date', ''),
                record.get('to_date', ''),
                record.get('notes', ''),
                record.get('doctor_name', ''),
            ]
            
            for col_num, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        
        # Freeze top row
        ws.freeze_panes = 'A2'
        
        # Save workbook
        output_filename = f"rekap_surat_sakit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = TEMPS_DIR / output_filename
        wb.save(str(output_path))
        
        logger.info(f"Excel file generated successfully: {output_path}")
        return str(output_path)
        
    except Exception as e:
        logger.error(f"Error generating Excel file: {str(e)}")
        raise
